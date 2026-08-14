#!/usr/bin/env python3
"""Build a candidate content package from a reviewed Excel workbook.

Workbook sheets: character_base, child_content, experience.
List cells use `|` separators. This tool writes JSON and a deterministic manifest;
it never copies media or changes the Android catalog.
"""
from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook

from compile_layered_content import LayerError, compile_package, validate_layers


REQUIRED = {
    "character_base": {"id", "character", "unicode", "pinyin", "tone", "strokeCount", "source", "sourceLicense", "sourceSnapshot"},
    "child_content": {"characterId", "meaningForChild", "teachingPrompt", "words", "sentence", "learningGoal", "imageTheme", "imageRequest", "audioRequest", "reviewStatus"},
    "experience": {"characterId", "learningOrder", "questionSeeds", "rewardProfile", "reviewProfile", "unlockAfter"},
}


def rows(workbook, sheet_name: str) -> list[dict]:
    if sheet_name not in workbook.sheetnames:
        raise LayerError(f"Excel 缺少工作表: {sheet_name}")
    sheet = workbook[sheet_name]
    values = list(sheet.values)
    if not values:
        raise LayerError(f"工作表为空: {sheet_name}")
    headers = [str(value).strip() if value is not None else "" for value in values[0]]
    missing = REQUIRED[sheet_name] - set(headers)
    if missing:
        raise LayerError(f"{sheet_name} 缺少列: {', '.join(sorted(missing))}")
    result = []
    for row in values[1:]:
        if all(value is None or str(value).strip() == "" for value in row):
            continue
        item = {header: value for header, value in zip(headers, row) if header}
        result.append(item)
    if not result:
        raise LayerError(f"工作表没有数据: {sheet_name}")
    return result


def clean(records: list[dict], list_fields: set[str]) -> list[dict]:
    cleaned = []
    for record in records:
        item = dict(record)
        for field in list_fields:
            value = item.get(field, "")
            item[field] = [part.strip() for part in str(value).split("|") if part and part.strip()]
        for field in ("tone", "strokeCount", "frequency", "learningOrder"):
            if field in item and item[field] not in (None, ""):
                item[field] = int(item[field])
        cleaned.append(item)
    return cleaned


def write_json(path: Path, value: object) -> None:
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--version", default="2.0.0-candidate")
    parser.add_argument("--allow-draft", action="store_true")
    args = parser.parse_args()
    try:
        workbook = load_workbook(args.workbook, read_only=True, data_only=True)
        base = clean(rows(workbook, "character_base"), set())
        child = clean(rows(workbook, "child_content"), {"words", "misconceptions"})
        experience = clean(rows(workbook, "experience"), {"questionSeeds", "unlockAfter"})
        child_by_id = {item["characterId"]: item for item in child}
        experience_by_id = {item["characterId"]: item for item in experience}
        if len(child_by_id) != len(child) or len(experience_by_id) != len(experience):
            raise LayerError("儿童内容或体验层存在重复 characterId")
        compiled = []
        for foundation in base:
            character_id = foundation["id"]
            if character_id not in child_by_id or character_id not in experience_by_id:
                raise LayerError(f"{character_id} 缺少儿童内容或体验记录")
            child_record, experience_record = child_by_id[character_id], experience_by_id[character_id]
            validate_layers(foundation, child_record, experience_record, args.allow_draft)
            compiled.append(compile_package(foundation, child_record, experience_record, args.version))
        characters = sorted((item["characters"][0] for item in compiled), key=lambda item: item["order"])
        if [item["order"] for item in characters] != list(range(1, len(characters) + 1)):
            raise LayerError("learningOrder 必须从 1 连续编号")
        package = compiled[0]
        package["learningOrder"] = [item["id"] for item in characters]
        package["characters"] = characters
        package["optionCatalog"] = [{"id": f"text_{item['id']}", "kind": "TEXT", "characterId": item["id"], "text": item["character"]} for item in characters]
        package["childLearningPack"]["characters"] = [entry for item in compiled for entry in item["childLearningPack"]["characters"]]
        args.output.mkdir(parents=True, exist_ok=True)
        content_path = args.output / "content.json"
        manifest_path = args.output / "manifest.json"
        write_json(content_path, package)
        payload = content_path.read_bytes()
        write_json(manifest_path, {
            "manifestVersion": 1,
            "status": "CANDIDATE",
            "contentVersion": args.version,
            "resources": [{"path": "content.json", "sha256": hashlib.sha256(payload).hexdigest(), "bytes": len(payload), "required": True}],
        })
        print(f"Candidate package written: {args.output} ({len(characters)} characters)")
    except (OSError, ValueError, LayerError) as exc:
        raise SystemExit(f"CONTENT BUILD ERROR: {exc}")


if __name__ == "__main__":
    main()
